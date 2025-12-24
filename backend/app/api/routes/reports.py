from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import datetime, timedelta
import io
import csv

from app.core.database import get_db
from app.core.auth import get_current_user
from app.models.models import (
    User, Product, Inventory, Warehouse, SalesOrder, SalesOrderItem,
    Category, Supplier, InventoryAlert, OrderStatus, InventoryTransaction,
    TransactionType
)
from app.schemas.schemas import (
    InventoryValueReport,
    SalesSummaryReport,
    SalesSummaryReport,
    ProductPerformance,
    DetailedSalesReport
)

router = APIRouter()


@router.get("/inventory-valuation", response_model=List[InventoryValueReport])
async def get_inventory_valuation(
    warehouse_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get inventory valuation report by warehouse."""
    query = db.query(
        Warehouse.id.label("warehouse_id"),
        Warehouse.name.label("warehouse_name"),
        func.count(func.distinct(Product.id)).label("total_products"),
        func.sum(Inventory.quantity_on_hand).label("total_quantity"),
        func.sum(Inventory.quantity_on_hand * Product.cost_price).label("total_value")
    ).join(
        Inventory, Warehouse.id == Inventory.warehouse_id
    ).join(
        Product, Inventory.product_id == Product.id
    )
    
    if warehouse_id:
        query = query.filter(Warehouse.id == warehouse_id)
    
    query = query.group_by(Warehouse.id, Warehouse.name)
    
    results = query.all()
    
    return [
        InventoryValueReport(
            warehouse_id=r.warehouse_id,
            warehouse_name=r.warehouse_name,
            total_products=r.total_products or 0,
            total_quantity=r.total_quantity or 0,
            total_value=r.total_value or 0.0
        )
        for r in results
    ]


@router.get("/sales-summary", response_model=SalesSummaryReport)
async def get_sales_summary(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get sales summary report for a date range."""
    # Default to last 30 days if no dates provided
    if not end_date:
        end_date_dt = datetime.now()
    else:
        end_date_dt = datetime.fromisoformat(end_date)
    
    if not start_date:
        start_date_dt = end_date_dt - timedelta(days=30)
    else:
        start_date_dt = datetime.fromisoformat(start_date)
    
    query = db.query(
        func.count(SalesOrder.id).label("total_orders"),
        func.sum(SalesOrder.total_amount).label("total_revenue"),
        func.sum(SalesOrderItem.quantity).label("total_items_sold")
    ).join(
        SalesOrderItem, SalesOrder.id == SalesOrderItem.sales_order_id
    ).filter(
        SalesOrder.order_date.between(start_date_dt, end_date_dt),
        SalesOrder.status != OrderStatus.CANCELLED
    )
    
    result = query.first()
    
    total_orders = result.total_orders or 0
    total_revenue = result.total_revenue or 0.0
    total_items_sold = result.total_items_sold or 0
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0.0
    
    return SalesSummaryReport(
        total_orders=total_orders,
        total_revenue=total_revenue,
        total_items_sold=total_items_sold,
        average_order_value=avg_order_value
    )


@router.get("/product-performance", response_model=List[ProductPerformance])
async def get_product_performance(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = Query(10, ge=1, le=100),
    sort_by: str = Query("revenue", regex="^(revenue|quantity)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get top performing products by revenue or quantity sold."""
    # Default to last 30 days
    if not end_date:
        end_date_dt = datetime.now()
    else:
        end_date_dt = datetime.fromisoformat(end_date)
    
    if not start_date:
        start_date_dt = end_date_dt - timedelta(days=30)
    else:
        start_date_dt = datetime.fromisoformat(start_date)
    
    query = db.query(
        Product.id.label("product_id"),
        Product.name.label("product_name"),
        func.sum(SalesOrderItem.quantity).label("total_sold"),
        func.sum(SalesOrderItem.line_total).label("total_revenue")
    ).join(
        SalesOrderItem, Product.id == SalesOrderItem.product_id
    ).join(
        SalesOrder, SalesOrderItem.sales_order_id == SalesOrder.id
    ).filter(
        SalesOrder.order_date.between(start_date_dt, end_date_dt),
        SalesOrder.status != OrderStatus.CANCELLED
    ).group_by(
        Product.id, Product.name
    )
    
    if sort_by == "revenue":
        query = query.order_by(func.sum(SalesOrderItem.line_total).desc())
    else:
        query = query.order_by(func.sum(SalesOrderItem.quantity).desc())
    
    results = query.limit(limit).all()
    
    return [
        ProductPerformance(
            product_id=r.product_id,
            product_name=r.product_name,
            total_sold=r.total_sold or 0,
            total_revenue=r.total_revenue or 0.0
        )
        for r in results
    ]


@router.get("/low-stock-summary")
async def get_low_stock_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get summary of products with low stock."""
    low_stock_products = db.query(
        Product, Inventory
    ).join(
        Inventory, Product.id == Inventory.product_id
    ).filter(
        Inventory.quantity_on_hand <= Product.reorder_point
    ).all()
    
    summary = {
        "total_low_stock_products": len(low_stock_products),
        "products": [
            {
                "product_id": p.id,
                "product_name": p.name,
                "sku": p.sku,
                "current_quantity": inv.quantity_on_hand,
                "reorder_point": p.reorder_point,
                "reorder_quantity": p.reorder_quantity,
                "warehouse_id": inv.warehouse_id
            }
            for p, inv in low_stock_products
        ]
    }
    
    return summary
"""GST Tax Report endpoint - append to reports.py"""

@router.get("/gst-summary")
async def get_gst_summary(
    start_date: str,
    end_date: str,
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate GST tax summary report grouped by tax rate
    
    Query params:
    - start_date: YYYY-MM-DD
    - end_date: YYYY-MM-DD
    - format: json | csv | excel
    """
    try:
        start_dt = datetime.fromisoformat(start_date)
        end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format. Use YYYY-MM-DD"
        )
    
    # Query sales order items grouped by tax_rate
    gst_data = db.query(
        SalesOrderItem.tax_rate,
        func.sum((SalesOrderItem.quantity * SalesOrderItem.unit_price) - SalesOrderItem.discount).label('taxable_amount'),
        func.sum(SalesOrderItem.tax_amount).label('tax_collected'),
        func.count(func.distinct(SalesOrderItem.sales_order_id)).label('order_count'),
        func.sum(SalesOrderItem.quantity).label('items_sold')
    ).join(
        SalesOrder, SalesOrderItem.sales_order_id == SalesOrder.id
    ).filter(
        SalesOrder.created_at >= start_dt,
        SalesOrder.created_at < end_dt,
        SalesOrder.status != 'cancelled'
    ).group_by(
        SalesOrderItem.tax_rate
    ).order_by(
        SalesOrderItem.tax_rate
    ).all()
    
    # Format data
    summary = []
    total_taxable = 0
    total_tax = 0
    total_orders = 0
    total_items = 0
    
    for row in gst_data:
        taxable = float(row.taxable_amount or 0)
        tax = float(row.tax_collected or 0)
        orders = int(row.order_count or 0)
        items = int(row.items_sold or 0)
        
        summary.append({
            "tax_rate": float(row.tax_rate or 0),
            "taxable_amount": taxable,
            "tax_collected": tax,
            "order_count": orders,
            "items_sold": items
        })
        
        total_taxable += taxable
        total_tax += tax
        total_orders += orders
        total_items += items
    
    report_data = {
        "period": {
            "start_date": start_date,
            "end_date": end_date
        },
        "summary": summary,
        "totals": {
            "taxable_amount": total_taxable,
            "tax_collected": total_tax,
            "order_count": total_orders,
            "items_sold": total_items
        }
    }
    
    # Return format based on request
    if format == "csv":
        return generate_gst_csv(report_data)
    elif format == "excel":
        return generate_gst_excel(report_data)
    else:
        return report_data


def generate_gst_csv(report_data: dict) -> StreamingResponse:
    """Generate CSV file from GST report data"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(['GST Tax Report'])
    writer.writerow(['Period', f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
    writer.writerow([])
    
    # Column headers
    writer.writerow(['GST Rate (%)', 'Taxable Amount (₹)', 'Tax Collected (₹)', 'Orders', 'Items Sold'])
    
    # Data rows
    for item in report_data['summary']:
        writer.writerow([
            f"{item['tax_rate']:.0f}",
            f"{item['taxable_amount']:.2f}",
            f"{item['tax_collected']:.2f}",
            item['order_count'],
            item['items_sold']
        ])
    
    # Total row
    writer.writerow([])
    totals = report_data['totals']
    writer.writerow([
        'Total',
        f"{totals['taxable_amount']:.2f}",
        f"{totals['tax_collected']:.2f}",
        totals['order_count'],
        totals['items_sold']
    ])
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=gst_report_{report_data['period']['start_date']}_to_{report_data['period']['end_date']}.csv"}
    )


def generate_gst_excel(report_data: dict) -> StreamingResponse:
    """Generate Excel file from GST report data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return generate_gst_csv(report_data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "GST Report"
    
    # Title
    ws['A1'] = 'GST Tax Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Period: {report_data['period']['start_date']} to {report_data['period']['end_date']}"
    
    # Headers
    headers = ['GST Rate (%)', 'Taxable Amount (₹)', 'Tax Collected (₹)', 'Orders', 'Items Sold']
    ws.append([])
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for item in report_data['summary']:
        ws.append([
            float(item['tax_rate']),
            float(item['taxable_amount']),
            float(item['tax_collected']),
            int(item['order_count']),
            int(item['items_sold'])
        ])
    
    # Total row
    ws.append([])
    totals = report_data['totals']
    total_row = ws.max_row + 1
    ws.append([
        'Total',
        float(totals['taxable_amount']),
        float(totals['tax_collected']),
        int(totals['order_count']),
        int(totals['items_sold'])
    ])
    
    # Style total row
    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=gst_report_{report_data['period']['start_date']}_to_{report_data['period']['end_date']}.xlsx"}
    )

@router.get("/detailed-sales-report")
async def get_detailed_sales_report(
    start_date: str,
    end_date: str,
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate Detailed Sales Report with Profit & Liability
    """
    try:
        start_dt = datetime.fromisoformat(start_date)
        end_dt = datetime.fromisoformat(end_date) + timedelta(days=1)
    except ValueError:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format"
        )

    # Query: SalesOrderItem joined with SalesOrder and Product
    results = db.query(
        SalesOrder.id.label("order_id"),
        SalesOrder.order_date,
        SalesOrder.order_number,
        SalesOrder.discount_amount.label("order_discount"),
        Product.name.label("product_name"),
        Product.cost_price.label("cost_price_unit"), # Current cost price
        SalesOrderItem.quantity,
        SalesOrderItem.unit_price.label("selling_price_unit"),
        SalesOrderItem.discount.label("item_discount"),
        SalesOrderItem.tax_rate,
        SalesOrderItem.tax_amount,
        SalesOrderItem.line_total
    ).join(
        SalesOrder, SalesOrderItem.sales_order_id == SalesOrder.id
    ).join(
        Product, SalesOrderItem.product_id == Product.id
    ).filter(
        SalesOrder.order_date >= start_dt,
        SalesOrder.order_date < end_dt,
        SalesOrder.status != OrderStatus.CANCELLED
    ).order_by(
        SalesOrder.order_date.desc()
    ).all()

    report_data = []
    
    total_line_profit_excl_gst = 0
    total_line_profit_inc_gst = 0
    total_gst_liability = 0
    
    seen_orders = set()
    total_order_discounts = 0

    for row in results:
        # Track order level discount once per order
        if row.order_id not in seen_orders:
            total_order_discounts += float(row.order_discount or 0)
            seen_orders.add(row.order_id)

        # Calculations
        qty = int(row.quantity or 0)
        
        # Cost Price
        cost_unit_excl_gst = float(row.cost_price_unit or 0)
        cost_total_excl_gst = cost_unit_excl_gst * qty
        
        # Calculate Cost Inc GST for display purposes
        tax_multiplier = 1 + ((float(row.tax_rate or 18)) / 100.0)
        cost_total_inc_gst = cost_total_excl_gst * tax_multiplier
        
        # Selling (Revenue)
        selling_gross = (float(row.selling_price_unit or 0) * qty)
        item_discount = float(row.item_discount or 0)
        selling_total_excl_gst = selling_gross - item_discount
        
        gst_liability = float(row.tax_amount or 0)
        selling_total_inc_gst = selling_total_excl_gst + gst_liability

        # Profit
        profit_excl_gst = selling_total_excl_gst - cost_total_excl_gst
        profit_inc_gst = selling_total_inc_gst - cost_total_excl_gst

        total_line_profit_excl_gst += profit_excl_gst
        total_line_profit_inc_gst += profit_inc_gst
        total_gst_liability += gst_liability

        report_data.append({
            "sale_date": row.order_date.strftime("%Y-%m-%d"),
            "order_number": row.order_number,
            "product_name": row.product_name,
            "quantity": qty,
            "item_discount": item_discount,
            "order_discount": float(row.order_discount or 0),
            "cost_total_excl_gst": cost_total_excl_gst,
            "cost_total_inc_gst": cost_total_inc_gst,
            "selling_total_excl_gst": selling_total_excl_gst,
            "selling_total_inc_gst": selling_total_inc_gst,
            "gst_liability": gst_liability,
            "profit_excl_gst": profit_excl_gst,
            "profit_inc_gst": profit_inc_gst
        })

    # Final Totals: Deduct order-level discounts from the summed line profits
    final_profit_excl_gst = total_line_profit_excl_gst - total_order_discounts
    final_profit_inc_gst = total_line_profit_inc_gst - total_order_discounts
    
    # Calculate sum of all discounts (line item + order level)
    sum_line_item_discounts = sum(float(item['item_discount']) for item in report_data)
    total_all_discounts = sum_line_item_discounts + total_order_discounts

    summary = {
        "period": {"start_date": start_date, "end_date": end_date},
        "items": report_data,
        "totals": {
            "profit_excl_gst": final_profit_excl_gst,
            "profit_inc_gst": final_profit_inc_gst,
            "gst_liability": total_gst_liability,
            "order_discounts": total_order_discounts,
            "total_all_discounts": total_all_discounts
        }
    }

    if format == "excel":
        return generate_detailed_sales_excel(summary)
    
    return summary


def generate_detailed_sales_excel(data: dict) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, 
            detail="openpyxl not installed"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Detailed Sales"

    # Headers
    headers = [
        "Sale Date", "Order #", "Product Name", "Qty", 
        "Item Disc.", "Order Disc.",
        "Cost (Excl GST)", "Cost (Inc GST)", 
        "Selling (Excl GST)", "Selling (Inc GST)", 
        "GST Liability", "Profit (Excl GST)", "Profit (Inc GST)"
    ]
    
    ws.append(["Detailed Sales Report", f"{data['period']['start_date']} to {data['period']['end_date']}"])
    ws.append([]) # spacer
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for item in data['items']:
        ws.append([
            item['sale_date'],
            item['order_number'],
            item['product_name'],
            item['quantity'],
            float(f"{item['item_discount']:.2f}"),
            float(f"{item['order_discount']:.2f}"),
            float(f"{item['cost_total_excl_gst']:.2f}"),
            float(f"{item['cost_total_inc_gst']:.2f}"),
            float(f"{item['selling_total_excl_gst']:.2f}"),
            float(f"{item['selling_total_inc_gst']:.2f}"),
            item['gst_liability'],
            item['profit_excl_gst'],
            item['profit_inc_gst']
        ])

    # Totals Row
    ws.append([])
    ws.append([
        "TOTALS", "", "", "", "", "", "", "", "", "",
        float(f"{data['totals']['gst_liability']:.2f}"),
        float(f"{data['totals']['profit_excl_gst']:.2f}"),
        float(f"{data['totals']['profit_inc_gst']:.2f}")
    ])
    
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=detailed_sales_report_{data['period']['start_date']}_to_{data['period']['end_date']}.xlsx"}
    )


@router.get("/stock-inventory")
async def get_stock_inventory_report(
    date: Optional[str] = Query(None),
    format: str = "json",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get Stock Inventory Report (Real-time or Historical)
    """
    now = datetime.now()
    if not date:
        target_date = now
    else:
        try:
            # Handle both YYYY-MM-DD and full ISO format
            if len(date) == 10:
                target_date = datetime.fromisoformat(date).replace(hour=23, minute=59, second=59)
            else:
                target_date = datetime.fromisoformat(date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    # Real-time is today or future
    is_real_time = target_date >= now.replace(hour=0, minute=0, second=0)

    if is_real_time:
        # Pull from Inventory table for "True Real Time"
        results = db.query(
            Product.name.label("product_name"),
            Product.sku,
            Warehouse.name.label("warehouse_name"),
            Inventory.quantity_on_hand,
            Inventory.quantity_reserved,
            Product.cost_price
        ).join(
            Inventory, Product.id == Inventory.product_id
        ).join(
            Warehouse, Inventory.warehouse_id == Warehouse.id
        ).all()
        
        report_items = [
            {
                "product_name": r.product_name,
                "sku": r.sku,
                "warehouse_name": r.warehouse_name,
                "quantity_on_hand": r.quantity_on_hand,
                "quantity_reserved": r.quantity_reserved,
                "available_quantity": r.quantity_on_hand - r.quantity_reserved,
                "valuation": r.quantity_on_hand * r.cost_price
            }
            for r in results
        ]
    else:
        # BACKTRACK LOGIC: 
        # Historical Qty = Current Qty - (Transactions since target_date)
        
        # 1. Get current inventory
        current_inv = db.query(
            Product.id.label("product_id"),
            Product.name.label("product_name"),
            Product.sku,
            Warehouse.id.label("warehouse_id"),
            Warehouse.name.label("warehouse_name"),
            Inventory.quantity_on_hand.label("current_qty"),
            Product.cost_price
        ).join(
            Inventory, Product.id == Inventory.product_id
        ).join(
            Warehouse, Inventory.warehouse_id == Warehouse.id
        ).all()

        report_items = []
        for r in current_inv:
            # 2. Get transactions that happened BETWEEN target_date AND now
            # We subtract these from current inventory to "go back in time"
            trans_since = db.query(
                func.sum(InventoryTransaction.quantity)
            ).filter(
                InventoryTransaction.product_id == r.product_id,
                InventoryTransaction.warehouse_id == r.warehouse_id,
                InventoryTransaction.created_at > target_date
            ).scalar() or 0

            historical_qty = int(r.current_qty) - int(trans_since)
            
            report_items.append({
                "product_name": r.product_name,
                "sku": r.sku,
                "warehouse_name": r.warehouse_name,
                "quantity_on_hand": historical_qty,
                "quantity_reserved": 0, # Historical reserved state not stored
                "available_quantity": historical_qty,
                "valuation": float(historical_qty) * float(r.cost_price or 0)
            })

    summary = {
        "report_date": target_date.strftime("%Y-%m-%d"),
        "is_real_time": is_real_time,
        "items": report_items,
        "totals": {
            "total_quantity": sum(item["quantity_on_hand"] for item in report_items),
            "total_valuation": sum(item["valuation"] for item in report_items)
        }
    }

    if format == "excel":
        return generate_stock_inventory_excel(summary)
    
    return summary


def generate_stock_inventory_excel(data: dict) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Inventory"

    # Headers
    headers = ["Product Name", "SKU", "Warehouse", "Qty On Hand", "Reserved", "Available", "Valuation (₹)"]
    ws.append(["Stock Inventory Report", f"As of: {data['report_date']}"])
    ws.append([])
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for item in data['items']:
        ws.append([
            item['product_name'],
            item['sku'],
            item['warehouse_name'],
            item['quantity_on_hand'],
            item['quantity_reserved'],
            item['available_quantity'],
            float(f"{item['valuation']:.2f}")
        ])

    # Totals Row
    ws.append([])
    ws.append([
        "TOTALS", "", "", 
        data['totals']['total_quantity'], 
        "", "",
        float(f"{data['totals']['total_valuation']:.2f}")
    ])
    
    # Auto-width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=stock_inventory_{data['report_date']}.xlsx"}
    )
